#!/usr/bin/env python3
"""
Left-hand punch feature analysis (Jab vs Lead Hook vs Lead Uppercut).

Loads V4-V10 MotionBERT 3D pose data + annotations, extracts per-clip
biomechanical features for each left-hand punch, then dumps:
  - Per-class mean/std table (sorted by ANOVA F-score)
  - Top discriminating features with interpretation
  - Intra-class consistency (CV) for each class
  - punch_features.csv  — full per-clip feature matrix

Usage:
    python punch_feature_analysis.py
    python punch_feature_analysis.py --out my_features.csv
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats
from scipy.signal import savgol_filter

from preprocessing import poses_to_body_frame_clip

# ── Paths ─────────────────────────────────────────────────────────────────────
REPO     = Path(__file__).parent.resolve()
POSE_DIR = REPO / "MotionBERT_3d"
ANN_DIR  = REPO / "Annotation_files"

VERSIONS = [f"V{i}" for i in range(4, 11)]

# V7 confirmed from mp4 that matches npy frame count; V9/V10 from short preview mp4s
FPS_MAP: dict[str, float] = {"V7": 29.97, "V9": 23.98, "V10": 25.0}

# ── Labels ────────────────────────────────────────────────────────────────────
_LABEL_NORM = {
    "jab":            "Jab",
    "cross":          "Cross",
    "lead hook":      "Lead Hook",
    "rear hook":      "Rear Hook",
    "lead uppercut":  "Lead Uppercut",
    "rear uppercut":  "Rear Uppercut",
}
LEFT_CLASSES = {"Jab", "Lead Hook", "Lead Uppercut"}
CLASS_ORDER  = ["Jab", "Lead Hook", "Lead Uppercut"]

def _norm_label(raw: str) -> str:
    return _LABEL_NORM.get(raw.strip().lower(), raw.strip())

# ── H36M joint indices ────────────────────────────────────────────────────────
_PELVIS  = 0
_R_HIP   = 1
_L_HIP   = 4
_THORAX  = 8
_L_SH    = 11
_L_EL    = 12
_L_WR    = 13
_R_SH    = 14
_R_EL    = 15
_R_WR    = 16

# ── Annotation loading ────────────────────────────────────────────────────────
def _load_annotations(xlsx: Path) -> list[tuple[int, int, str]]:
    """
    Returns list of (start_1idx, end_1idx, label).

    Handles the three xlsx layouts across V4-V10:
      dense  (V4-V8):  cols 0=start, 1=end, 2=label
      sparse (V9-V10): cols 0=start, 2=end, 4=label  (None spacers at 1,3)
    """
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    is_sparse = len(header) > 4 and header[1] is None and header[2] is not None
    sc, ec, lc = (0, 2, 4) if is_sparse else (0, 1, 2)

    rows: list[tuple[int, int, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= lc:
            continue
        s, e, lab = row[sc], row[ec], row[lc]
        if s is None or e is None or lab is None:
            continue
        try:
            s, e = int(s), int(e)
        except (TypeError, ValueError):
            continue
        if e >= s:
            rows.append((s, e, str(lab).strip()))
    wb.close()
    return rows

def _smooth(arr: np.ndarray, win: int = 5) -> np.ndarray:
    """Savitzky-Golay smoothing along time axis."""
    T = arr.shape[0]
    if T < win:
        return arr
    flat = arr.reshape(T, -1)
    s = savgol_filter(flat, window_length=win, polyorder=min(2, win - 1), axis=0)
    return s.reshape(arr.shape)

# ── Angle helpers ─────────────────────────────────────────────────────────────
def _angle3(a: np.ndarray, b: np.ndarray, c: np.ndarray) -> float:
    """Interior angle at b (degrees)."""
    ba, bc = a - b, c - b
    denom = np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-8
    return float(np.degrees(np.arccos(np.clip(np.dot(ba, bc) / denom, -1.0, 1.0))))


def _arm_elevation(frame: np.ndarray, side: str = "left") -> float:
    """
    Angle between the upper arm and the hanging direction (-Z).
    0° = arm at rest, 90° = horizontal, 180° = overhead.
    """
    sh, el = (_L_SH, _L_EL) if side == "left" else (_R_SH, _R_EL)
    arm = frame[el] - frame[sh]
    n = np.linalg.norm(arm) + 1e-8
    cos_t = -arm[2] / n
    return float(np.degrees(np.arccos(np.clip(cos_t, -1.0, 1.0))))

# ── Per-clip feature extraction ───────────────────────────────────────────────
def extract_features(clip_raw: np.ndarray, fps: float | None) -> dict | None:
    """
    Extract scalar features from one left-hand punch clip.

    Parameters
    ----------
    clip_raw : (T, 17, 3) raw H36M poses from X3D.npy
    fps      : frames-per-second; None → skip speed-magnitude features

    Returns
    -------
    dict of scalar features, or None if clip is too short.
    """
    T = clip_raw.shape[0]
    if T < 3:
        return None

    q = _smooth(poses_to_body_frame_clip(clip_raw.astype(np.float64)))  # (T,17,3)

    # ── Left-arm elbow angles across clip ─────────────────────────────────
    elbow = np.array([_angle3(q[t, _L_SH], q[t, _L_EL], q[t, _L_WR]) for t in range(T)])

    # ── Left-arm elevation across clip ────────────────────────────────────
    elev = np.array([_arm_elevation(q[t], "left") for t in range(T)])

    # ── Left wrist-to-shoulder distance (reach) ───────────────────────────
    reach = np.linalg.norm(q[:, _L_WR] - q[:, _L_SH], axis=-1)

    # ── Left wrist position in body frame ─────────────────────────────────
    wrist = q[:, _L_WR]   # (T,3)  X=lateral, Y=forward, Z=vertical

    # Extension frame = frame of maximum elbow angle (most stable reference)
    ext_t = int(np.argmax(elbow))

    # ── Wrist displacement: start → extension frame ────────────────────────
    disp = wrist[ext_t] - wrist[0]           # (3,) in body frame
    disp_y = float(disp[1])                  # forward  → jab
    disp_x = float(abs(disp[0]))             # lateral  → hook
    disp_z = float(disp[2])                  # vertical → uppercut

    # Trajectory angles (body-frame geometry)
    # punch_angle_yz: arctan2(z, y)  >0 = upward  (uppercut signal)
    # punch_angle_xy: arctan2(|x|,y) >0 = lateral (hook signal)
    angle_yz = float(np.degrees(np.arctan2(disp[2], disp[1] + 1e-8)))
    angle_xy = float(np.degrees(np.arctan2(abs(disp[0]), abs(disp[1]) + 1e-8)))

    # ── Forearm unit vector at extension (wrist − elbow) ──────────────────
    fa = q[ext_t, _L_WR] - q[ext_t, _L_EL]
    fa_mag = np.linalg.norm(fa) + 1e-8
    fa_n = fa / fa_mag
    fa_abs = np.abs(fa_n)
    # Forearm elevation: arcsin(z-component) in degrees
    fa_elev = float(np.degrees(np.arcsin(np.clip(fa_n[2], -1.0, 1.0))))

    # ── Velocity direction at extension frame ──────────────────────────────
    if ext_t > 0:
        vel = wrist[ext_t] - wrist[ext_t - 1]
    elif T > 1:
        vel = wrist[1] - wrist[0]
    else:
        vel = np.zeros(3)
    vel_mag = np.linalg.norm(vel) + 1e-8
    vel_abs = np.abs(vel)

    # ── Wrist speed (frame-to-frame displacement magnitude) ───────────────
    wrist_diff = np.linalg.norm(np.diff(wrist, axis=0), axis=-1)  # (T-1,)
    peak_speed_frames = float(wrist_diff.max()) if len(wrist_diff) > 0 else 0.0
    peak_speed = peak_speed_frames * fps if fps is not None else None

    # ── Shoulder and hip yaw ──────────────────────────────────────────────
    sh_vec  = q[:, _R_SH] - q[:, _L_SH]    # (T,3)
    hip_vec = q[:, _R_HIP] - q[:, _L_HIP]
    sh_yaw  = np.degrees(np.arctan2(sh_vec[:, 1],  sh_vec[:, 0]  + 1e-8))
    hip_yaw = np.degrees(np.arctan2(hip_vec[:, 1], hip_vec[:, 0] + 1e-8))
    xfactor = sh_yaw - hip_yaw  # shoulder rotation relative to hips

    # ── Guard (right) arm elbow ───────────────────────────────────────────
    guard_elbow = np.array([
        _angle3(q[t, _R_SH], q[t, _R_EL], q[t, _R_WR]) for t in range(T)
    ])

    feats: dict = {
        # Elbow angle (left arm)
        "elbow_peak":       float(elbow.max()),
        "elbow_at_ext":     float(elbow[ext_t]),
        "elbow_range":      float(elbow.max() - elbow.min()),
        "elbow_min":        float(elbow.min()),

        # Arm elevation (left arm, 0=hanging, 90=horiz)
        "elev_peak":        float(elev.max()),
        "elev_at_ext":      float(elev[ext_t]),
        "elev_range":       float(elev.max() - elev.min()),

        # Reach (wrist-shoulder distance in torso-lengths)
        "reach_peak":       float(reach.max()),
        "reach_at_ext":     float(reach[ext_t]),
        "reach_range":      float(reach.max() - reach.min()),

        # Wrist trajectory (start → extension frame)
        "wrist_y_disp":     disp_y,    # forward displacement  → jab
        "wrist_x_disp":     disp_x,    # lateral displacement  → hook
        "wrist_z_disp":     disp_z,    # vertical displacement → uppercut
        "punch_angle_yz":   angle_yz,  # >0 = upward  (uppercut)
        "punch_angle_xy":   angle_xy,  # >0 = lateral (hook)

        # Forearm direction at extension (unit vector fractions)
        "fa_X_frac":        float(fa_abs[0]),    # lateral  → hook
        "fa_Y_frac":        float(fa_abs[1]),    # forward  → jab
        "fa_Z_frac":        float(fa_abs[2]),    # vertical → uppercut
        "fa_elev_deg":      fa_elev,             # >0 = angled upward → uppercut

        # Velocity direction fractions at extension (FPS-independent)
        "vel_X_frac":       float(vel_abs[0] / vel_mag),
        "vel_Y_frac":       float(vel_abs[1] / vel_mag),
        "vel_Z_frac":       float(vel_abs[2] / vel_mag),

        # Wrist height (body-frame Z in torso-lengths)
        "wrist_z_at_ext":   float(wrist[ext_t, 2]),
        "wrist_z_peak":     float(wrist[:, 2].max()),
        "wrist_z_start":    float(wrist[0, 2]),

        # Shoulder/hip rotation
        "sh_yaw_range":     float(sh_yaw.max() - sh_yaw.min()),
        "hip_yaw_range":    float(hip_yaw.max() - hip_yaw.min()),
        "xfactor_at_ext":   float(xfactor[ext_t]),
        "xfactor_max":      float(xfactor.max()),
        "xfactor_range":    float(xfactor.max() - xfactor.min()),

        # Guard arm
        "guard_elbow_min":  float(guard_elbow.min()),
        "guard_elbow_mean": float(guard_elbow.mean()),

        # Speed (frame-count units; fps-scaled where known)
        "peak_speed_frames": peak_speed_frames,

        # Clip duration
        "clip_frames":      float(T),
    }

    if peak_speed is not None:
        feats["peak_speed"] = peak_speed

    return feats


# ── Statistics helpers ────────────────────────────────────────────────────────
def _anova_rank(df: pd.DataFrame, feature_cols: list[str]) -> list[tuple[str, float, float]]:
    """One-way ANOVA across the three left-hand classes. Returns sorted (col, F, p) list."""
    results = []
    for col in feature_cols:
        groups = [df[df["label"] == cls][col].dropna().values for cls in CLASS_ORDER]
        groups = [g for g in groups if len(g) > 1]
        if len(groups) < 2:
            continue
        try:
            F, p = stats.f_oneway(*groups)
            if np.isfinite(F):
                results.append((col, float(F), float(p)))
        except Exception:
            pass
    results.sort(key=lambda x: -x[1])
    return results


def _print_separator(width: int = 110) -> None:
    print("  " + "─" * width)


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="Left-hand punch feature analysis")
    parser.add_argument("--out", default="punch_features.csv",
                        help="Output CSV path (default: punch_features.csv)")
    args = parser.parse_args()

    records: list[dict] = []

    print("\nLoading punch clips...")
    for ver in VERSIONS:
        npy_path = POSE_DIR / ver / "X3D.npy"
        ann_path = ANN_DIR / f"{ver}.xlsx"
        if not npy_path.exists() or not ann_path.exists():
            print(f"  [skip] {ver}: missing file")
            continue

        fps = FPS_MAP.get(ver)
        frames  = np.load(npy_path)
        n_total = frames.shape[0]
        anns    = _load_annotations(ann_path)

        kept, skipped = 0, 0
        for s_1idx, e_1idx, raw_label in anns:
            label = _norm_label(raw_label)
            if label not in LEFT_CLASSES:
                continue
            # 1-indexed → 0-indexed slice
            s0 = s_1idx - 1
            e0 = min(e_1idx, n_total)
            if e0 <= s0:
                skipped += 1
                continue
            clip = frames[s0:e0]
            feats = extract_features(clip, fps)
            if feats is None:
                skipped += 1
                continue
            feats["label"]   = label
            feats["version"] = ver
            records.append(feats)
            kept += 1

        fps_str = f"{fps:.2f}" if fps else "unknown"
        print(f"  {ver}: {kept} clips  (fps={fps_str})")

    if not records:
        print("No records found. Check data paths.")
        sys.exit(1)

    df = pd.DataFrame(records)

    out_path = REPO / args.out
    df.to_csv(out_path, index=False)

    # ── Summary counts ────────────────────────────────────────────────────────
    total = len(df)
    counts = {cls: int((df["label"] == cls).sum()) for cls in CLASS_ORDER}
    print(f"\nTotal: {total} clips  ({', '.join(f'{c}={n}' for c, n in counts.items())})")
    print(f"Saved feature matrix → {out_path}\n")

    # ── ANOVA ranking ─────────────────────────────────────────────────────────
    feature_cols = [c for c in df.columns if c not in ("label", "version")]
    anova = _anova_rank(df, feature_cols)
    anova_dict = {col: (F, p) for col, F, p in anova}

    # ── Feature interpretation guide ──────────────────────────────────────────
    INTERP = {
        "punch_angle_yz":    "↑ = upward trajectory (uppercut), ↓ = forward (jab/hook)",
        "punch_angle_xy":    "↑ = lateral trajectory (hook), ↓ = straight (jab/uppercut)",
        "wrist_z_disp":      "↑ = wrist rises (uppercut), ↓ = level/falling",
        "wrist_y_disp":      "↑ = wrist drives forward (jab)",
        "wrist_x_disp":      "↑ = wrist moves laterally (hook)",
        "elev_peak":         "↑ = arm raised higher (jab/hook), ↓ = low (uppercut)",
        "elev_at_ext":       "↑ = arm elevation at extension",
        "reach_peak":        "↑ = arm more extended (jab)",
        "reach_at_ext":      "↑ = arm extension at peak-elbow frame",
        "elbow_peak":        "↑ = straighter arm (jab), ↓ = bent (hook/uppercut)",
        "elbow_at_ext":      "same; measured at peak-elbow frame",
        "elbow_range":       "↑ = more elbow movement through clip",
        "fa_Y_frac":         "↑ = forearm pointing forward (jab)",
        "fa_Z_frac":         "↑ = forearm pointing upward (uppercut)",
        "fa_X_frac":         "↑ = forearm pointing laterally (hook)",
        "fa_elev_deg":       "↑ = forearm angled upward (uppercut), ↓ = horizontal",
        "vel_Y_frac":        "↑ = wrist moving forward at ext (jab)",
        "vel_Z_frac":        "↑ = wrist moving upward at ext (uppercut)",
        "vel_X_frac":        "↑ = wrist moving laterally at ext (hook)",
        "wrist_z_at_ext":    "↑ = wrist position higher (jab/hook)",
        "wrist_z_peak":      "↑ = highest wrist Z in clip",
        "sh_yaw_range":      "↑ = more shoulder rotation through clip",
        "xfactor_max":       "↑ = more shoulder-hip separation (torsion)",
        "xfactor_range":     "↑ = more shoulder-hip rotation range",
        "guard_elbow_min":   "↑ = guard arm more extended",
        "peak_speed_frames": "↑ = faster wrist (in frame-count units)",
        "peak_speed":        "↑ = faster wrist (torso-L/s, where fps known)",
        "clip_frames":       "clip duration in frames",
    }

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 1: All features sorted by ANOVA F-score
    # ─────────────────────────────────────────────────────────────────────────
    print("=" * 112)
    print(f"  FEATURE TABLE  —  sorted by ANOVA F-score  (higher = better class separation)")
    print(f"  Classes: {', '.join(f'{c} (n={counts[c]})' for c in CLASS_ORDER)}")
    print("=" * 112)

    col_w = 22
    hdr = (f"  {'Feature':<{col_w}}"
           f"  {'Jab mean':>10}  {'Jab std':>8}"
           f"  {'Hook mean':>10}  {'Hook std':>8}"
           f"  {'Upcut mean':>10}  {'Upcut std':>8}"
           f"  {'F':>8}  {'p':>8}")
    print(hdr)
    _print_separator()

    for col, F, p in anova:
        parts = [f"  {col:<{col_w}}"]
        for cls in CLASS_ORDER:
            sub = df[df["label"] == cls][col].dropna()
            m = sub.mean() if len(sub) > 0 else float("nan")
            s = sub.std()  if len(sub) > 0 else float("nan")
            parts.append(f"  {m:>10.3f}  {s:>8.3f}")
        parts.append(f"  {F:>8.2f}  {p:>8.4f}")
        print("".join(parts))

    print()

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 2: Top 15 discriminating features with interpretation
    # ─────────────────────────────────────────────────────────────────────────
    print("=" * 112)
    print("  TOP 15 DISCRIMINATING FEATURES")
    print("=" * 112)
    print(f"  {'Rank':<5}  {'Feature':<{col_w}}  {'F':>8}  {'p':>8}  Interpretation")
    _print_separator()

    for rank, (col, F, p) in enumerate(anova[:15], 1):
        interp = INTERP.get(col, "")
        print(f"  {rank:<5}  {col:<{col_w}}  {F:>8.2f}  {p:>8.4f}  {interp}")

    print()

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 3: Per-class intra-class consistency (lowest CV = most consistent)
    # ─────────────────────────────────────────────────────────────────────────
    print("=" * 112)
    print("  INTRA-CLASS CONSISTENCY  —  top 12 most consistent features per class")
    print("  (CV = std / |mean|, lower = more consistent across same-class clips)")
    print("=" * 112)

    for cls in CLASS_ORDER:
        sub = df[df["label"] == cls][feature_cols].dropna(axis=1, how="all")
        means = sub.mean()
        stds  = sub.std()
        cvs   = (stds / (means.abs() + 1e-8))
        cvs   = cvs.replace([np.inf, -np.inf], np.nan).dropna().sort_values()

        n_cls = counts[cls]
        print(f"\n  {cls}  (n={n_cls})")
        print(f"  {'Feature':<{col_w}}  {'mean':>9}  {'std':>9}  {'CV':>7}")
        print("  " + "-" * 55)
        for feat, cv in cvs.head(12).items():
            m = means[feat]
            s = stds[feat]
            print(f"  {feat:<{col_w}}  {m:>9.3f}  {s:>9.3f}  {cv:>7.3f}")

    print()

    # ─────────────────────────────────────────────────────────────────────────
    # TABLE 4: Pairwise separability (t-test between each class pair)
    # ─────────────────────────────────────────────────────────────────────────
    print("=" * 112)
    print("  PAIRWISE SEPARABILITY  —  top 10 features per class pair  (by |t-statistic|)")
    print("=" * 112)

    pairs = [
        ("Jab", "Lead Hook",      "Jab vs Hook"),
        ("Jab", "Lead Uppercut",  "Jab vs Uppercut"),
        ("Lead Hook", "Lead Uppercut", "Hook vs Uppercut"),
    ]

    for cls_a, cls_b, pair_label in pairs:
        a_data = df[df["label"] == cls_a]
        b_data = df[df["label"] == cls_b]
        tresults = []
        for col in feature_cols:
            ga = a_data[col].dropna().values
            gb = b_data[col].dropna().values
            if len(ga) < 2 or len(gb) < 2:
                continue
            try:
                t, p = stats.ttest_ind(ga, gb, equal_var=False)
                if np.isfinite(t):
                    tresults.append((col, float(t), float(p),
                                     ga.mean(), gb.mean()))
            except Exception:
                pass
        tresults.sort(key=lambda x: -abs(x[1]))

        print(f"\n  {pair_label}  "
              f"(n={counts[cls_a]} vs n={counts[cls_b]})")
        print(f"  {'Feature':<{col_w}}  {'t-stat':>8}  {'p':>8}  "
              f"  {cls_a[:8]:>10} mean  {cls_b[:10]:>12} mean")
        print("  " + "-" * 80)
        for col, t, p, ma, mb in tresults[:10]:
            print(f"  {col:<{col_w}}  {t:>8.2f}  {p:>8.4f}  {ma:>14.3f}  {mb:>16.3f}")

    print()

    # ─────────────────────────────────────────────────────────────────────────
    # Summary: best single-feature rules
    # ─────────────────────────────────────────────────────────────────────────
    print("=" * 112)
    print("  CANDIDATE SINGLE-FEATURE DECISION RULES")
    print("=" * 112)
    print()

    # Find best threshold for punch_angle_yz to separate uppercut
    if "punch_angle_yz" in df.columns:
        upcut_vals = df[df.label == "Lead Uppercut"]["punch_angle_yz"].dropna()
        other_vals = df[df.label != "Lead Uppercut"]["punch_angle_yz"].dropna()
        if len(upcut_vals) > 0 and len(other_vals) > 0:
            thresh = (upcut_vals.mean() + other_vals.mean()) / 2
            print(f"  Uppercut detector (punch_angle_yz):")
            print(f"    Jab mean={df[df.label=='Jab']['punch_angle_yz'].mean():.1f}°  "
                  f"Hook mean={df[df.label=='Lead Hook']['punch_angle_yz'].mean():.1f}°  "
                  f"Uppercut mean={upcut_vals.mean():.1f}°")
            print(f"    Suggested threshold: punch_angle_yz > {thresh:.1f}° → Uppercut")

    if "reach_peak" in df.columns:
        jab_r = df[df.label == "Jab"]["reach_peak"].dropna()
        hook_r = df[df.label == "Lead Hook"]["reach_peak"].dropna()
        if len(jab_r) > 0 and len(hook_r) > 0:
            thresh_jab = (jab_r.mean() + hook_r.mean()) / 2
            print(f"\n  Jab vs Hook (reach_peak after uppercut filter):")
            print(f"    Jab mean={jab_r.mean():.3f}  "
                  f"Hook mean={hook_r.mean():.3f}")
            print(f"    Suggested threshold: reach_peak >= {thresh_jab:.3f} → Jab")

    if "fa_Z_frac" in df.columns:
        upcut_z = df[df.label == "Lead Uppercut"]["fa_Z_frac"].dropna()
        other_z = df[df.label != "Lead Uppercut"]["fa_Z_frac"].dropna()
        if len(upcut_z) > 0 and len(other_z) > 0:
            thresh_z = (upcut_z.mean() + other_z.mean()) / 2
            print(f"\n  Forearm vertical fraction (fa_Z_frac) — alternative uppercut signal:")
            print(f"    Jab mean={df[df.label=='Jab']['fa_Z_frac'].mean():.3f}  "
                  f"Hook mean={df[df.label=='Lead Hook']['fa_Z_frac'].mean():.3f}  "
                  f"Uppercut mean={upcut_z.mean():.3f}")
            print(f"    Suggested threshold: fa_Z_frac > {thresh_z:.3f} → Uppercut")

    print()
    print("=" * 112)
    print(f"  Full feature matrix saved to: {out_path}")
    print("=" * 112)
    print()


if __name__ == "__main__":
    main()
